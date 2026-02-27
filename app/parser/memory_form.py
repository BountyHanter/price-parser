from typing import Any
from openpyxl import load_workbook

from app.parser.schema import Task

NAME_COL = 2  # Наименование (вторая колонка)


def build_memory_structure(xlsx_path: str, tasks: list[Task]) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    rows: dict[str, Any] = {}

    for t in tasks:
        row_key = str(t.row)

        item = rows.get(row_key)
        if item is None:
            name_val = ws.cell(row=t.row, column=NAME_COL).value
            name = "" if name_val is None else str(name_val).strip()

            item = {"name": name, "sites": {}}
            rows[row_key] = item

        item["sites"][t.site_name] = {
            "url_col": t.url_col,
            "price_col": t.price_col,
            "url": t.url,
            "price": None,
            "error": None,
        }

    return {"rows": rows}
import json

from openpyxl import load_workbook
from pathlib import Path

from app.fast_api_logger import log
from app.parser.paths import JSON_PATH, XLSX_PATH, OUTPUT_PATH

# сколько изменений делать перед промежуточным сохранением
BATCH_SIZE = 1000


def load_data(path: str | Path) -> dict:
    path = Path(path)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_prices():
    try:
        log.info("[WRITE] Начинаю запись цен в Excel")

        data = load_data(JSON_PATH)

        wb = load_workbook(XLSX_PATH)
        ws = wb.active

        rows = data["rows"]

        counter = 0
        total_written = 0

        for row_str, row_data in rows.items():
            row_index = int(row_str)
            sites = row_data.get("sites", {})

            for site_name, site_data in sites.items():
                price = site_data.get("price")
                price_col = site_data.get("price_col")

                if price is None or price_col is None:
                    continue

                ws.cell(row=row_index, column=price_col, value=price)

                counter += 1
                total_written += 1

                if counter >= BATCH_SIZE:
                    log.info(f"[WRITE] batch save ({total_written} cells)")
                    wb.save(OUTPUT_PATH)
                    counter = 0

        if OUTPUT_PATH.exists():
            OUTPUT_PATH.unlink()

        wb.save(OUTPUT_PATH)
        log.info(f"[WRITE] Готово. Записано ячеек: {total_written}")

    except Exception:
        log.exception("[WRITE] Ошибка при записи Excel")
        raise

if __name__ == "__main__":
    write_prices()
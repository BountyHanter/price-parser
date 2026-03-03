import json
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.parser.memory_form import build_memory_structure
from app.parser.schema import Task, SITES


def _is_url(value) -> bool:
    if value is None:
        return False
    return bool(str(value).strip())


def normalize_url(url: str) -> str:
    url = url.strip()

    if not url:
        return url

    if url.startswith(("http://", "https://")):
        return url

    return f"https://{url}"


def read_tasks_from_excel(
    xlsx_path: str | Path,
    sheet_name: str | None = None,
    start_row: int = 3,
) -> list[Task]:

    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    tasks: list[Task] = []

    for row in range(start_row, ws.max_row + 1):

        for site_name, cols in SITES.items():
            url_col = cols["url_col"]
            price_col = cols["price_col"]


            val = ws.cell(row=row, column=url_col).value
            if not _is_url(val):
                continue

            url = normalize_url(str(val))

            tasks.append(
                Task(
                    row=row,
                    site_name=site_name,
                    url_col=url_col,
                    price_col=price_col,
                    url=url,
                    price_cell=f"{get_column_letter(price_col)}{row}"
                )
            )

    return tasks


if __name__ == "__main__":
    tasks = read_tasks_from_excel("template.xlsx")
    print(len(tasks))
    print(tasks[:5])
    result = build_memory_structure("template.xlsx", tasks)
    print(json.dumps(result, indent=2, ensure_ascii=False))